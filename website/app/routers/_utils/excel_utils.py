import io
from openpyxl import load_workbook
from ...model.pacient_hospital import models, schemas, old_models
from sqlalchemy import func, and_, or_

# Функция для удаления временной зоны и преобразования в строки только с датой для всех остальных столбцов

def adjust_column_widths(workbook):
    """
    Автоматически настраивает ширину всех столбцов для каждого листа в рабочей книге.

    Параметры:
    workbook (openpyxl.workbook.workbook.Workbook): Рабочая книга, в которой настраиваются ширины столбцов.

    Описание:
    Функция проходит по всем листам рабочей книги и устанавливает ширину столбцов в зависимости от 
    максимальной длины значения в каждой ячейке. Ширина столбцов устанавливается с учетом фиксированного 
    значения и некоторого запаса.
    """
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Получаем букву столбца
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            max_length = 3 / 0.14  # Константная ширина
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

def apply_filters_to_excel(output):
    """
    Применяет фильтры ко всем листам в рабочей книге и настраивает ширину столбцов.

    Параметры:
    output (io.BytesIO): Объект BytesIO, содержащий байты рабочей книги Excel.

    Возвращает:
    io.BytesIO: Обновленный объект BytesIO с примененными фильтрами и отрегулированными ширинами столбцов.

    Описание:
    Функция загружает рабочую книгу из байтового потока, применяет фильтры ко всем диапазонам данных 
    на каждом листе и затем сохраняет обновленную рабочую книгу обратно в байтовый поток.
    """
    workbook = load_workbook(filename=io.BytesIO(output.getvalue()))
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet.auto_filter.ref = sheet.dimensions  # Применение фильтра ко всему диапазону данных
    
    adjust_column_widths(workbook)  # Автоматическая настройка ширины столбцов

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

def get_latest_patient_movements(db):
    """
    Получает последние события движения пациентов из таблицы PatientMovement.

    Параметры:
    db (sqlalchemy.orm.session.Session): Объект сессии SQLAlchemy для выполнения запросов к базе данных.

    Возвращает:
    list of dict: Список словарей, где каждый словарь представляет собой запись о движении пациента с полями из таблицы PatientMovement.

    Описание:
    Функция выполняет запрос к базе данных для получения последних событий для каждого пациента, основываясь 
    на самой последней дате события. Результаты преобразуются в список словарей, где каждый словарь содержит 
    данные о пациенте.
    """
    latest_movements_subquery = (
        db.query(
            models.PatientMovement.patient_id,
            func.max(models.PatientMovement.event_date).label('latest_event_date')
        )
        .group_by(models.PatientMovement.patient_id)
        .subquery()
    )

    latest_movements = (
        db.query(models.PatientMovement)
        .join(
            latest_movements_subquery,
            models.PatientMovement.patient_id == latest_movements_subquery.c.patient_id
        )
        .filter(
            models.PatientMovement.event_date == latest_movements_subquery.c.latest_event_date
        )
        .all()
    )

    # Преобразование данных SQLAlchemy в словари
    latest_movements_data = [
        {column.name: getattr(movement, column.name) for column in models.PatientMovement.__table__.columns}
        for movement in latest_movements
    ]

    return latest_movements_data
