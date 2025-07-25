from fastapi import APIRouter, BackgroundTasks, HTTPException, Depends, Request, Form, Query, Body, status
from fastapi.responses import HTMLResponse
from fastapi import HTTPException
from starlette.authentication import SimpleUser, requires
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi import FastAPI, Request, Depends, Form, HTTPException
from sqlalchemy.orm import Session, joinedload 
from typing import List, Tuple, Optional, Dict
from ...model.database.database import SessionLocal
from ...model.сlinic.models import ClinicData as ClinicDataModel
from ...model.сlinic.models import DoctorsOfTheClinic, ClinicAppointment, VvkClinic, HospitalizationClinic
from ...model.сlinic.schemas import ClinicData, ClinicDataCreate, ClinicAppointmentCreate, PatientUpdate, VvkCreate, Vvk, HospitalizationCreate, HospitalizationData, AppointmentUpdate
from ...model.сlinic.schemas import DoctorsOfTheClinic as DoctorsOfTheClinicSchema
from ...model.сlinic.schemas import ClinicAppointment as ClinicAppointmentSchema
from ...model.сlinic.schemas import UpdatePatientData
from fastapi.responses import JSONResponse
from datetime import datetime, timedelta
from sqlalchemy import case, distinct, func, and_
from fastapi.templating import Jinja2Templates
import os

router = APIRouter()

current_dir = os.path.dirname(os.path.abspath(__file__))
docs_dir = os.path.abspath(os.path.join(current_dir, "../../static/docs"))
img_dir = os.path.abspath(os.path.join(current_dir, "../static/img"))
js_dir = os.path.abspath(os.path.join(current_dir, "../static/js"))
css_dir = os.path.abspath(os.path.join(current_dir, "../static/css"))
templates = Jinja2Templates(directory=os.path.abspath(os.path.join(current_dir, "../../templates")))

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

def remove_timezone(dt):
    if isinstance(dt, datetime) and dt.tzinfo is not None:
        return dt.replace(tzinfo=None)
    return dt

def format_datetime(value):
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    return value

def format_date(value):
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return value

import threading

@router.get("/download_excel/")
def download_excel_clinic(db: Session = Depends(get_db)):
    # Получаем текущую дату и время
    now = datetime.now()

    # Форматируем дату и время в нужный формат
    formatted_date_time = now.strftime("%d_%m_%Y_%H_%M")

    # Создаем новое название файла
    file_name = f"Учёт амбулаторных пациентов поликлиника на {formatted_date_time}.xlsx"
    try:
        print("Starting report generation...")

        wb = Workbook()
        headers = {
            'Амбулаторные карты': ['ID', 'Номер карты', 'ФИО',  'Направление', 'День рождения', 'Адрес',
                            'Документ', 'Воинское подразделение', 'Воинское звание', 'Принадлежность',
                            'Боевые действия', 'Личный номер', 'Военный комиссариат', 'Военный округ',
                            'Номер телефона', 'Информация о пациенте', 'Медицинский тип',
                            'Текущее время'],
            'Врачи поликлиники': ['ID', 'Тип доктора', 'ФИО', 'Кабинет', 'Номер телефона'],
            'Записи к врачам': ['ID', 'ФИО пациента', 'ФИО доктора', 'Дата и время', 'Примечание',
                                'Отчет доктора', 'Заметка доктора'],
            'Направления на госпитализацию': ['ID', 'ФИО пациента', 'Дата госпитализации', 'Диагноз', 'Направление'],
            'ВВК': ['ID', 'ФИО пациента', 'Дата VVK', 'Номер VVK', 'Заключение']
        }

        try:
            clinic_data = db.query(ClinicDataModel).all()
            doctors = db.query(DoctorsOfTheClinic).all()
            appointments = db.query(ClinicAppointment).all()
            hospitalizations = db.query(HospitalizationClinic).all()
            vvks = db.query(VvkClinic).all()
            print("Data retrieved from database")
        except Exception as e:
            print(f"Database query failed: {str(e)}")
            raise HTTPException(status_code=500, detail="Error retrieving data from the database")

        # Create a dictionary for quick lookup of patient names by ID
        patient_name_lookup = {data.id: data.full_name for data in clinic_data}
        doctor_name_lookup = {doctor.id: doctor.full_name for doctor in doctors}

        # Define fixed width for columns
        fixed_column_width = 20

        for sheet_name, cols in headers.items():
            ws = wb.create_sheet(title=sheet_name.capitalize())
            ws.append(cols)

            if sheet_name == 'Амбулаторные карты':
                for data in clinic_data:
                    row = [
                        data.id,
                        data.card_number,
                        data.full_name,
                        data.directed,
                        format_date(remove_timezone(data.birthday)) if data.birthday else None,
                        data.address,
                    
                        data.document,
                        data.military_unit,
                        data.military_rank,
                        data.belonging,
                        data.hostilities,
                        data.personal_number,
                        data.military_commissariat,
                        data.military_district,
                        data.phone_number,
                        data.patient_info,
                        data.medical_type,
                        format_datetime(remove_timezone(data.current_time)) if data.current_time else None
                    ]
                    ws.append(row)

            elif sheet_name == 'Врачи поликлиники':
                for data in doctors:
                    row = [data.id, data.doctor_type, data.full_name, data.cabinet, data.phone_number]
                    ws.append(row)

            elif sheet_name == 'Записи к врачам':
                for data in appointments:
                    row = [
                        data.id,
                        patient_name_lookup.get(data.patient_id, "Unknown"),
                        doctor_name_lookup.get(data.doctor_id, "Unknown"),
                        format_datetime(remove_timezone(data.datetime_q)) if data.datetime_q else "",
                        data.note,
                        data.doctor_report,
                        data.doctor_note
                    ]
                    ws.append(row)

            elif sheet_name == 'Направления на госпитализацию':
                for data in hospitalizations:
                    row = [
                        data.id,
                        patient_name_lookup.get(data.patient_id, "Unknown"),
                        format_date(remove_timezone(data.date)) if data.date else "",
                        data.diagnosis,
                        data.direction
                    ]
                    ws.append(row)

            elif sheet_name == 'ВВК':
                for data in vvks:
                    row = [
                        data.id,
                        patient_name_lookup.get(data.patient_id, "Unknown"),
                        format_date(remove_timezone(data.vvk_date)) if data.vvk_date else "",
                        data.vvk_number,
                        data.conclusion
                    ]
                    ws.append(row)

            # Set fixed width for all columns
            for col in ws.columns:
                for cell in col:
                    col_letter = get_column_letter(cell.column)
                    ws.column_dimensions[col_letter].width = fixed_column_width

        # Save the workbook
        wb.save(file_name)

        def delete_file_after_delay(file_path, delay):
            import time
            time.sleep(delay)
            if os.path.exists(file_path):
                os.remove(file_path)
                print(f"File {file_path} deleted.")

        # Start a thread to delete the file after 5 minutes (300 seconds)
        threading.Thread(target=delete_file_after_delay, args=(file_name, 300)).start()

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating Excel report: {str(e)}")

    return FileResponse(path=file_name, filename=file_name, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

from pytz import timezone
moscow_tz = timezone('Europe/Moscow')

@router.get("/hospitalizations/by-date")
def hospitalizations_by_date_page(request: Request, db: Session = Depends(get_db), date: str = Query(None)):
    if date:
        selected_date = datetime.strptime(date, '%Y-%m-%d').date()
    else:
        selected_date = datetime.now(moscow_tz).date()

    # Получаем госпитализации пациентов на выбранную дату
    hospitalizations = (
        db.query(HospitalizationClinic)
        .join(HospitalizationClinic.patient)
        .options(joinedload(HospitalizationClinic.patient))  # Подгружаем информацию о пациенте
        .filter(func.date(HospitalizationClinic.date) == selected_date)
        .all()
    )

    # Преобразуем строку в datetime для поля birthday, если это необходимо
    for hospitalization in hospitalizations:
        if isinstance(hospitalization.patient.birthday, str):
            try:
                hospitalization.patient.birthday = datetime.strptime(hospitalization.patient.birthday, '%Y-%m-%d')
            except ValueError:
                hospitalization.patient.birthday = None  # Если формат даты некорректен

    return templates.TemplateResponse("clinic/hospitalizations_by_date.html", {
        "request": request,
        "hospitalizations": hospitalizations,
        "selected_date": selected_date.strftime('%Y-%m-%d'),
        "message": "Госпитализации не найдены" if not hospitalizations else ""
    })

@router.get("/by-date")
def patients_by_date_page(request: Request, db: Session = Depends(get_db), date: str = Query(None)):
    if date:
        selected_date = datetime.strptime(date, '%Y-%m-%d').date()
    else:
        selected_date = datetime.now(moscow_tz).date()

    # Получаем всех пациентов, созданных на выбранную дату
    patients = (
        db.query(ClinicDataModel)
        .options(joinedload(ClinicDataModel.appointments).joinedload(ClinicAppointment.doctor))
        .filter(func.date(ClinicDataModel.current_time) == selected_date)
        .all()
    )

    # Преобразуем строку в datetime для поля birthday, если это необходимо
    for patient in patients:
        if isinstance(patient.birthday, str):
            try:
                patient.birthday = datetime.strptime(patient.birthday, '%Y-%m-%d')
            except ValueError:
                patient.birthday = None  # Если формат даты некорректен

    return templates.TemplateResponse("clinic/patients_today.html", {
        "request": request,
        "patients": patients,
        "selected_date": selected_date.strftime('%Y-%m-%d'),
        "message": "Пациенты не найдены" if not patients else ""
    })


#region vvk
@router.post("/vvk/", response_model=VvkCreate)
def create_vvk(vvk: VvkCreate, db: Session = Depends(get_db)):
    try:
        print(vvk.dict())  # Выводим содержимое словаря vvk.dict()
        new_vvk = VvkClinic(**vvk.dict())
        print(vvk.dict())  # Повторно выводим для проверки
        db.add(new_vvk)
        db.commit()
        db.refresh(new_vvk)
        return new_vvk
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating VVK: {str(e)}")
    
@router.get("/vvk/", response_model=List[Vvk])
def get_vvks(patient_id: int = Query(...), db: Session = Depends(get_db)):
    vvks = db.query(VvkClinic).filter(VvkClinic.patient_id == patient_id).all()
    return vvks

@router.delete("/vvk/{vvk_id}/", response_model=Vvk)
def delete_vvk(vvk_id: int, db: Session = Depends(get_db)):
    vvk = db.query(VvkClinic).get(vvk_id)
    if not vvk:
        raise HTTPException(status_code=404, detail="VVK entry not found")
    db.delete(vvk)
    db.commit()
    return vvk
#endregion

#region  hospitalization
@router.post("/patients/{patient_id}/hospitalizations/", response_model=HospitalizationCreate)
def create_hospitalization(patient_id: int, hospitalization: HospitalizationCreate, db: Session = Depends(get_db)):
    patient = db.query(ClinicDataModel).filter(ClinicDataModel.id == patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Пациент не найден")
    
    new_hospitalization = HospitalizationClinic(
        patient_id=patient_id,
        date=hospitalization.date,
        diagnosis=hospitalization.diagnosis,
        direction=hospitalization.direction
    )
    db.add(new_hospitalization)
    db.commit()
    db.refresh(new_hospitalization)
    return new_hospitalization

@router.get("/patients/{patient_id}/hospitalizations/", response_model=List[HospitalizationData])
def get_hospitalizations(patient_id: int, db: Session = Depends(get_db)):
    hospitalizations = db.query(HospitalizationClinic).filter(HospitalizationClinic.patient_id == patient_id).all()
    return hospitalizations

@router.delete("/hospitalizations/{hospitalization_id}/", response_model=dict)
def delete_hospitalization(hospitalization_id: int, db: Session = Depends(get_db)):
    hospitalization = db.query(HospitalizationClinic).filter(HospitalizationClinic.id == hospitalization_id).first()
    if not hospitalization:
        raise HTTPException(status_code=404, detail="Госпитализация не найдена")
    
    db.delete(hospitalization)
    db.commit()
    return {"Сообщение": "Информация о удалении направления на госпитлизацию успешно удалена."}
#endregion


# region doctors


@router.post("/doctors/add", response_class=HTMLResponse)
async def add_doctor(
    request: Request, 
    doctor_type: str = Form(...), 
    full_name: str = Form(...), 
    cabinet: str = Form(None), 
    phone_number: str = Form(None), 
    db: Session = Depends(get_db)
):
    new_doctor = DoctorsOfTheClinic(
        doctor_type=doctor_type,
        full_name=full_name,
        cabinet=cabinet,
        phone_number=phone_number
    )
    db.add(new_doctor)
    db.commit()
    return RedirectResponse(url="/clinic/doctors/administration", status_code=303)


@router.get("/doctors/delete/{id}", response_class=HTMLResponse)
async def delete_doctor(id: int, request: Request, db: Session = Depends(get_db)):
    doctor = db.query(DoctorsOfTheClinic).filter(DoctorsOfTheClinic.id == id).first()
    if doctor is None:
        raise HTTPException(status_code=404, detail="Doctor not found")
    db.delete(doctor)
    db.commit()
    return RedirectResponse(url="/clinic/doctors/administration", status_code=303)


@router.get("/doctors/administration", response_class=HTMLResponse)
async def get_doctor_list_admin(request: Request, db: Session = Depends(get_db)):
    doctors = db.query(DoctorsOfTheClinic).all()
    return templates.TemplateResponse("clinic/doctor_admin.html", {"request": request, "doctors": doctors})



@router.get("/doctors", response_class=HTMLResponse)
def get_doctor_list(request: Request, db: Session = Depends(get_db)):
    doctors = db.query(DoctorsOfTheClinic).all()
    return templates.TemplateResponse("clinic/doctors_list.html", {"request": request, "doctors": doctors})


@router.post("/appointments/check", response_model=dict)
def check_appointments_for_doctor(
    doctor_id: int = Form(...), 
    datetime_str: str = Form(...), 
    db: Session = Depends(get_db)
):
    print("Получен запрос на проверку записи:")
    print(f"doctor_id: {doctor_id}, datetime: {datetime_str}")

    try:
        datetime_value = datetime.fromisoformat(datetime_str)
        fifteen_minutes_later = datetime_value + timedelta(minutes=15)

        existing_appointment = (
            db.query(ClinicAppointment)
            .filter(
                ClinicAppointment.doctor_id == doctor_id,
                ClinicAppointment.datetime.between(datetime_value, fifteen_minutes_later)
            )
            .first()
        )
        
        if existing_appointment:
            print("Запись на прием невозможна: врач занят.")
            return {"allowed": False}
        else:
            print("Запись на прием разрешена: врач свободен.")
            return {"allowed": True}
    except Exception as e:
        print("Ошибка при проверке записи на прием:", e)
        raise HTTPException(status_code=500, detail="Internal server error")
    
@router.get("/doctors/{doctor_id}/appointments", response_class=HTMLResponse)
def get_doctor_appointments(request: Request, doctor_id: int, date: str = None, db: Session = Depends(get_db)):
    doctor = db.query(DoctorsOfTheClinic).filter(DoctorsOfTheClinic.id == doctor_id).first()
    if not doctor:
        raise HTTPException(status_code=404, detail="Doctor not found")

    query = db.query(ClinicAppointment).options(joinedload(ClinicAppointment.patient)).filter(ClinicAppointment.doctor_id == doctor_id)
    if date:
        query = query.filter(func.date(ClinicAppointment.datetime_q) == date)
    
    appointments = query.all()
    
    active_appointments_by_date = {}
    past_appointments_by_date = {}
    
    current_datetime = datetime.now()

    for appointment in appointments:
        appointment_datetime_naive = appointment.datetime_q.replace(tzinfo=None)
        appointment_date = appointment.datetime_q.date()
        if appointment_datetime_naive >= current_datetime:
            if appointment_date not in active_appointments_by_date:
                active_appointments_by_date[appointment_date] = []
            active_appointments_by_date[appointment_date].append(appointment)
        else:    
            if appointment_date not in past_appointments_by_date:
                past_appointments_by_date[appointment_date] = []
            past_appointments_by_date[appointment_date].append(appointment)
    
    return templates.TemplateResponse("clinic/doctor_appointments.html", {
        "request": request, 
        "doctor": doctor, 
        "active_appointments_by_date": active_appointments_by_date, 
        "past_appointments_by_date": past_appointments_by_date
    })

@router.post("/doctors/{doctor_id}/appointments", response_model=ClinicAppointmentSchema)
def create_appointment(
    doctor_id: int, 
    datetime_q: datetime = Form(...), 
    note: str = Form(...), 
    doctor_report: str = Form(...), 
    doctor_note: str = Form(...), 
    db: Session = Depends(get_db)
):
    appointment = ClinicAppointment(
        doctor_id=doctor_id,
        datetime_q=datetime_q,
        note=note,
        doctor_report=doctor_report,
        doctor_note=doctor_note
    )
    db.add(appointment)
    db.commit()
    db.refresh(appointment)
    return ClinicAppointmentSchema.from_orm(appointment)

@router.delete("/doctors/delete/{doctor_id}")
def delete_doctor(doctor_id: int, db: Session = Depends(get_db)):
    doctor = db.query(DoctorsOfTheClinic).filter(DoctorsOfTheClinic.id == doctor_id).first()
    if not doctor:
        raise HTTPException(status_code=404, detail="Doctor not found")
    
    db.delete(doctor)
    db.commit()
    return {"message": "Doctor deleted successfully"}
#endregion



# region clinic_patient

@router.delete("/delete_patient/{patient_id}")
def delete_patient(patient_id: int, db: Session = Depends(get_db)):
    # Найти пациента по patient_id
    patient = db.query(ClinicDataModel).filter(ClinicDataModel.id == patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Пациент не найден")
    
    # Удалить все записи о приёмах, связанных с пациентом
    db.query(ClinicAppointment).filter(ClinicAppointment.patient_id == patient_id).delete()
    
    # Удалить саму запись о пациенте
    db.delete(patient)
    db.commit()
    
    return {"message": "Пациент и связанные с ним записи успешно удалены"}

@router.get("/patient/{patient_id}")
def patient_page(request: Request, patient_id: int):
    #return templates.TemplateResponse("clinic/clinic_patient.html", {"request": request, "patient_id": patient_id})
    return templates.TemplateResponse("clinic/clinic_patient.html", {"request": request, "patient_id": patient_id})


from .._utils.documents_utils import (
    create_header_with_doctor_and_shape,
    fill_template,
    set_paragraph_formatting,
    add_custom_run,
    add_custom_run_yellow,
    set_cell_border,
    set_cell_vertical_alignment,
    center_text_in_line,
    add_full_underlined_text,
    set_font_to_times_new_roman,
    fill_template_v_times,
    format_name,
    delete_file_with_delay,
    set_cell_border_v2,
    set_table_borders,
    set_font_and_spacing
)


@router.get("/vvk-report/{patient_id}")
async def generate_document(
    request: Request,
    patient_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db)
):
    # Извлечение данных пациента
    patient = db.query(ClinicDataModel).filter(ClinicDataModel.id == patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    # Извлечение данных VVK
    vvk_data = db.query(VvkClinic).filter(VvkClinic.patient_id == patient_id).first()
    if not vvk_data:
        raise HTTPException(status_code=404, detail="VVK data not found")

    template_path = os.path.join(docs_dir, "templates", "Заключение ВВК поликлиника шаблон.docx")

    if isinstance(patient.birthday, str):
    # Attempt to convert from string to datetime
        try:
            patient.birthday = datetime.strptime(patient.birthday, '%Y-%m-%d')  # Adjust this format to your needs
        except ValueError:
            patient.birthday = None  # Handle invalid string format

    # Пример заполнения контекста для документа VVK
    context = {
        'full_name': patient.full_name.upper(),
        'today': datetime.now().strftime('%d.%m.%Y'),
        'military_unit': patient.military_unit if patient.military_unit else '',
        'birthday': patient.birthday.strftime('%d.%m.%Y') if patient.birthday else '',
        'military_rank': patient.military_rank.lower() if patient.military_rank else '',
        'vvk_date': vvk_data.vvk_date.strftime('%d.%m.%Y') if vvk_data.vvk_date else '',
        'vvk_number': vvk_data.vvk_number,
        'conclusion': vvk_data.conclusion,

    }


    # Создание документа
    output_path = fill_template_v_times(template_path, context, patient.full_name, file_name='Заключение ВВК поликлиника шаблон.docx')
    download_link = request.url_for('download_report', filename=os.path.basename(output_path))
    
    # Добавление задачи на удаление файла через 5 минут
    background_tasks.add_task(delete_file_with_delay, output_path, 300)
    
    return JSONResponse(content={"message": "Document generated successfully.", "download_link": str(download_link)})

@router.get("/hospitalization-report/{hospitalization_id}")
async def generate_hospitalization_document(
    request: Request,
    hospitalization_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db)
):
    # Извлечение данных о госпитализации
    hospitalization = db.query(HospitalizationClinic).filter(HospitalizationClinic.patient_id == hospitalization_id).first()
    if not hospitalization:
        raise HTTPException(status_code=404, detail="Hospitalization not found")

    # Извлечение данных пациента
    patient = db.query(ClinicDataModel).filter(ClinicDataModel.id == hospitalization.patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")

    template_path = os.path.join(docs_dir, "templates", "Заключение о госпитализации шаблон.docx")

    if isinstance(patient.birthday, str):
        # Попытка преобразовать строку в datetime
        try:
            patient.birthday = datetime.strptime(patient.birthday, '%Y-%m-%d')  # Настройте этот формат по необходимости
        except ValueError:
            patient.birthday = None  # Обработка неверного формата строки

    # Пример заполнения контекста для документа о госпитализации
    context = {
        'full_name': patient.full_name.upper(),
        'today': datetime.now().strftime('%d.%m.%Y'),
        'military_unit': patient.military_unit if patient.military_unit else '',
        'birthday': patient.birthday.strftime('%d.%m.%Y') if patient.birthday else '',
        'military_rank': patient.military_rank.lower() if patient.military_rank else '',
        'hospitalization_date': hospitalization.date.strftime('%d.%m.%Y') if hospitalization.date else '',
        'diagnosis': hospitalization.diagnosis,
        'direction': hospitalization.direction,
        # Добавьте другие необходимые поля из hospitalization
    }

    # Создание документа
    output_path = fill_template_v_times(template_path, context, patient.full_name, file_name='Заключение о госпитализации шаблон.docx')
    download_link = request.url_for('download_report', filename=os.path.basename(output_path))
    
    # Добавление задачи на удаление файла через 5 минут
    background_tasks.add_task(delete_file_with_delay, output_path, 300)
    
    return JSONResponse(content={"message": "Document generated successfully.", "download_link": str(download_link)})


# Обработка запроса на скачивание файла
@router.get("/download_report/{filename}")
async def download_report(filename: str):
    file_path = os.path.join(docs_dir, filename)
    if not os.path.isfile(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(file_path)


@router.get("/clinic_data/", response_model=List[ClinicData])
def get_patients(
    name: str = Query(None, min_length=3),
    db: Session = Depends(get_db)
):
    query = db.query(ClinicDataModel).options(
        joinedload(ClinicDataModel.hospitalizations),
        joinedload(ClinicDataModel.vvks)
    )

    if name:
        query = query.filter(ClinicDataModel.full_name.ilike(f"%{name}%"))
    
    query = query.order_by(ClinicDataModel.current_time.desc())
    
    patients = query.all()


    return patients

@router.put("/update_patient/{patient_id}")
def update_patient(patient_id: int, patient_update: PatientUpdate, db: Session = Depends(get_db)):
    patient = db.query(ClinicDataModel).filter(ClinicDataModel.id == patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    
    for key, value in patient_update.dict(exclude_unset=True).items():
        setattr(patient, key, value)
    
    db.commit()
    db.refresh(patient)
    return patient

@router.get("/appointments/{patient_id}", response_model=List[ClinicAppointmentSchema])
def get_patient_appointments(patient_id: int, db: Session = Depends(get_db)):
    appointments = (
        db.query(ClinicAppointment)
        .join(DoctorsOfTheClinic, ClinicAppointment.doctor_id == DoctorsOfTheClinic.id)
        .filter(ClinicAppointment.patient_id == patient_id)
        .all()
    )
    # Преобразование поля datetime в нужный формат (дд.мм.гггг в чч:мм)
    for appointment in appointments:
        appointment.datetime_q = appointment.datetime_q.strftime("%d.%m.%Y в %H:%M")
        appointment.note = appointment.note or ""  # По умолчанию пустая строка, если значение None
        appointment.doctor_report = appointment.doctor_report or ""  # По умолчанию пустая строка, если значение None
        appointment.doctor_note = appointment.doctor_note or ""  # По умолчанию пустая строка, если значение None
        appointment.doctor_full_name = f"{appointment.doctor.doctor_type} / {appointment.doctor.full_name}"
    return appointments

from sqlalchemy import or_, Date, cast
from datetime import date

def parse_date(date_str: str) -> datetime:
    try:
        return datetime.strptime(date_str, "%d %m %y")
    except ValueError:
        return None

@router.get("/", response_class=HTMLResponse)
async def read_and_search_patients(request: Request, search: str = ""):
    db: Session = SessionLocal()

    name = ""
    birthday = None

    if search:
        search_parts = search.split()
        name = " ".join([part for part in search_parts if not part.replace('.', '').isdigit()])
        date_part = " ".join([part for part in search_parts if part.replace('.', '').isdigit()])
        
        if date_part:
            try:
                birthday = parse(date_part, dayfirst=True)
            except ValueError:
                pass

    query = db.query(ClinicData)
    if name:
        query = query.filter(ClinicData.full_name.ilike(f"%{name}%"))
    if birthday:
        query = query.filter(ClinicData.birthday == birthday.date())

    patients = query.all()

    return templates.TemplateResponse("patients.html", {"request": request, "patients": patients})

@router.get("/patient/{patient_id}/details", response_class=HTMLResponse)
async def get_patient_details(request: Request, patient_id: int):
    db: Session = SessionLocal()
    patient = db.query(ClinicDataModel).filter(ClinicDataModel.id == patient_id).first()
    
    if not patient:
        return HTMLResponse(status_code=404, content="Пациент не найден")

    appointments = db.query(ClinicAppointment).filter(ClinicAppointment.patient_id == patient_id).all()
    hospitalizations = db.query(HospitalizationClinic).filter(HospitalizationClinic.patient_id == patient_id).all()
    vvks = db.query(VvkClinic).filter(VvkClinic.patient_id == patient_id).all()

   
    if isinstance(patient.birthday, str):
        try:
            patient.birthday = datetime.strptime(patient.birthday, '%Y-%m-%d')
        except ValueError:
            patient.birthday = None

    return templates.TemplateResponse("clinic/patient_details_modal.html", {
        "request": request, 
        "patient": patient,
        "appointments": appointments,
        "hospitalizations": hospitalizations,
        "vvks": vvks
    })

from .clinic_window import (update_patient_router,
                            create_appointment_router,
                            create_hospitalization_router,
                            create_vvk_router,
                            delete_appointment_router,
                            delete_hospitalization_router,
                            delete_vvk_router)



@router.put("/patients/{patient_id}/update", response_model=UpdatePatientData)
async def update_patient(patient_id: int, patient_data: UpdatePatientData, db: Session = Depends(get_db)):
    return update_patient_router(patient_id, patient_data, db)

@router.post("/patients/{patient_id}/appointments", response_model=ClinicAppointmentSchema)
async def create_appointment(patient_id: int, appointment_data: dict, db: Session = Depends(get_db)):
    return create_appointment_router(patient_id, appointment_data, db)

@router.post("/patients/{patient_id}/hospitalizations", response_model=HospitalizationData)
async def create_hospitalization(patient_id: int, hospitalization_data: dict, db: Session = Depends(get_db)):
    return create_hospitalization_router(patient_id, hospitalization_data, db)

@router.post("/patients/{patient_id}/vvks", response_model=Vvk)
async def create_vvk(patient_id: int, vvk_data: dict, db: Session = Depends(get_db)):
    return create_vvk_router(patient_id, vvk_data, db)

@router.delete("/appointments/{appointment_id}", status_code=204)
async def delete_appointment(appointment_id: int, db: Session = Depends(get_db)):
    return delete_appointment_router(appointment_id, db)

@router.delete("/hospitalizations/{hospitalization_id}", status_code=204)
async def delete_hospitalization(hospitalization_id: int, db: Session = Depends(get_db)):
    return delete_hospitalization_router(hospitalization_id, db)

@router.delete("/vvks/{vvk_id}", status_code=204)
async def delete_vvk(vvk_id: int, db: Session = Depends(get_db)):
    return delete_vvk_router(vvk_id, db)

# тут

@router.post("/appointments/", response_model=List[ClinicAppointmentSchema])
def add_appointment(
    patient_id: int = Form(...),
    doctor_id: int = Form(...),
    datetime_q: str = Form(...),  # Получаем строку из формы и затем преобразуем в datetime
    note: str = Form(None),
    doctor_report: str = Form(None),
    doctor_note: str = Form(None),
    db: Session = Depends(get_db)
):
    try:
        # Convert datetime string to datetime object
        datetime_obj = datetime.fromisoformat(datetime_q)
        
        # Get the doctor object from the database
        doctor = db.query(DoctorsOfTheClinic).filter(DoctorsOfTheClinic.id == doctor_id).first()
        
        # Check if the doctor exists
        if not doctor:
            raise HTTPException(status_code=404, detail="Doctor not found")

        # Create a new appointment object
        new_appointment = ClinicAppointment(
            patient_id=patient_id,
            doctor_id=doctor_id,
            datetime_q=datetime_obj,
            note=note,
            doctor_report=doctor_report or "",
            doctor_note=doctor_note or ""
        )

        # Add the new appointment to the database
        db.add(new_appointment)
        db.commit()

        return []

    except Exception as e:
        db.rollback()
        print("Database Error:", e)
        raise HTTPException(status_code=500, detail="Database error occurred")

@router.put("/doctors/{doctor_id}/appointments/{appointment_id}/update")
def update_appointment(
    doctor_id: int,
    appointment_id: int,
    appointment_update: AppointmentUpdate,
    db: Session = Depends(get_db)
):
    try:

        appointment = db.query(ClinicAppointment).filter(ClinicAppointment.id == appointment_id).first()
        
        if not appointment:
            raise HTTPException(status_code=404, detail="Appointment not found")

        appointment.doctor_report = appointment_update.doctor_report
        appointment.doctor_note = appointment_update.doctor_note
        
        db.commit()
        return {"message": "Appointment updated successfully", "appointment_id": appointment.id}
    except Exception as e:
        print(f"Error occurred: {e}")
        raise HTTPException(status_code=422, detail="Invalid data provided")



@router.get("/doctors/", response_model=List[DoctorsOfTheClinicSchema])
def get_doctors(db: Session = Depends(get_db)):
    doctors = db.query(DoctorsOfTheClinic).all()
    return doctors

@router.get("/clinic_data/{patient_id}", response_model=ClinicData)
def get_patient(patient_id: int, db: Session = Depends(get_db)):
    patient = db.query(ClinicDataModel).filter_by(id=patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    return patient


import urllib.parse
import logging
@router.get("/clinic_data/search/", response_model=List[ClinicData])
def search_patients(query: str = Query(..., min_length=3), db: Session = Depends(get_db)):
    if query is None or len(query) < 3:
        return []  # Возвращаем пустой список вместо ошибки 422

    try:
        decoded_query = urllib.parse.unquote(query)
        patients = db.query(ClinicDataModel).filter(
            (ClinicDataModel.full_name.contains(decoded_query)) | 
            (ClinicDataModel.card_number.contains(decoded_query))
        ).options(joinedload(ClinicDataModel.hospitalizations), joinedload(ClinicDataModel.vvks)).all()
        

        return patients
    except Exception as e:
        raise HTTPException(status_code=500, detail="Internal server error")

@router.post("/patient/", response_model=ClinicDataCreate)
def add_patient(
            full_name: str = Form(...),
            card_number: str = Form(...),
            birthday: str = Form(...),
            address: str = Form(None),
            belonging: str = Form(None),
            military_district: str = Form(None),
            military_commissariat: str = Form(None),
            hostilities: bool = Form(False),
            personal_number: str = Form(None),
            document: str = Form(None),
            phone_number: str = Form(None),
            military_unit: str = Form(None),
            military_rank: str = Form(None),
            directed: str = Form(...),
            db: Session = Depends(get_db)):

    existing_patient = db.query(ClinicDataModel).filter_by(full_name=full_name, birthday=birthday).first()
    if existing_patient:
        raise HTTPException(status_code=400, detail="Пациент уже существует")
    
    patient = ClinicDataModel(
        full_name=full_name,
        card_number=card_number,
        birthday=birthday,
        address=address,
        belonging=belonging,
        military_district=military_district,
        military_commissariat=military_commissariat,
        hostilities=hostilities,
        personal_number=personal_number,
        document=document,
        phone_number=phone_number,
        military_unit=military_unit,
        military_rank=military_rank,
        directed=directed
    )
    db.add(patient)
    db.commit()
    return RedirectResponse(url="/clinic/patient", status_code=302)


@router.get("/patient/")
def clinic_page(request: Request, db: Session = Depends(get_db)):
    # Получаем последнюю запись по номеру карты
    last_card = db.query(ClinicDataModel.card_number).order_by(ClinicDataModel.id.desc()).first()
    
    # Получаем общее количество пациентов
    total_patients = db.query(ClinicDataModel).count()
    
    

    # Передаем последний номер карты и общее количество пациентов в шаблон
    return templates.TemplateResponse("clinic/clinic_table.html", {
        "request": request,
        "last_card_number": last_card[0] if last_card else None,
        "total_patients": total_patients
    })

#endregion
@router.get("/statistics")
def get_statistics_page(request: Request):
    return templates.TemplateResponse("clinic/clinic_statistic.html", {"request": request})

@router.get("/statistics/monthly_patients")
def get_monthly_patients(db: Session = Depends(get_db)):
    result = db.query(
        func.to_char(ClinicAppointment.datetime_q, 'YYYY-MM').label('month'),
        func.count(ClinicAppointment.id).label('count'),
        func.count(ClinicAppointment.patient_id.distinct()).label('unique_patients')
    ).group_by('month').order_by('month').all()
    
    return [{'month': row.month, 'count': row.count} for row in result]


@router.get("/statistics/doctor_patients")
def get_doctor_patients(db: Session = Depends(get_db)):
    result = db.query(
        DoctorsOfTheClinic.full_name.label('doctor_name'),
        func.count(ClinicAppointment.id).label('count')
    ).join(ClinicAppointment, ClinicAppointment.doctor_id == DoctorsOfTheClinic.id).group_by(DoctorsOfTheClinic.full_name).all()
    return [{'doctor_name': row.doctor_name, 'count': row.count} for row in result]

@router.get("/statistics/overall")
def get_overall_statistics(db: Session = Depends(get_db)):
    today = date.today()

    total_appointments = db.query(func.count(ClinicAppointment.id)).scalar()

    total_patients = db.query(func.count(ClinicDataModel.id)).scalar()
    today_patients = db.query(func.count(ClinicAppointment.id)).filter(
        func.date(ClinicAppointment.datetime_q) == today
    ).scalar()
    vvk_today = db.query(func.count(VvkClinic.id)).filter(
        VvkClinic.vvk_date == today
    ).scalar()
    hospitalizations_today = db.query(func.count(HospitalizationClinic.id)).filter(
        HospitalizationClinic.date == today
    ).scalar()

    doctor_stats = db.query(
        DoctorsOfTheClinic.full_name.label('doctor_name'),
        func.count(ClinicAppointment.id).label('count')
    ).join(ClinicAppointment, ClinicAppointment.doctor_id == DoctorsOfTheClinic.id).group_by(DoctorsOfTheClinic.full_name).all()

    return {
        'total_patients': total_patients,
        'total_patients_today': today_patients,  # Изменено для соответствия клиентскому коду
        'vvk_today': vvk_today,
        'hospitalizations_today': hospitalizations_today,
        'doctor_stats': today_patients,
        'total_appointments': total_appointments
    }

